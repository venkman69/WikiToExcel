'''
Wiki To Excel converter

@copyright: Narayan Natarajan <venkman69@yahoo.com>
@author: venkman69
@license:
The MIT License (MIT)

Copyright (c) <year> <copyright holders>

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in
all copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
THE SOFTWARE.
'''
from copy import deepcopy
import re

from openpyxl import Workbook
from openpyxl.styles import fills
from openpyxl.styles.colors import Color
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.fonts import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.utils import coordinate_from_string
from __builtin__ import file
import os
from wikitoexcel.wikitblparser import wikiTableParser

HTML_BR=re.compile(r"<br[ ]*[/]>", re.IGNORECASE)

# import getpass
def getHTMLStyle(htmlNode):
    pass

def captionToExcel(capNode):
    # this is worksheet name
    return capNode.strip()

def procStyle(bsNode):        
    """proc style only retrieves the following style items:
    - font size
    - font weight (bold)
    - font color
    - font strikethrough
    - font underline
    - font italics
    - background color
    """
    procstyles=['font-weight',
                'font_name',
                'italic',
                'line-through',
                'underline',
                'background-color',
                'width',
                'color']
    styleMap=dict((k,False) for k in procstyles)
    for k,v in  bsNode.attrs.iteritems():
        if k == "font-family":
            styleMap['font_name']=v
        if k == "font-weight":
            if v=="bold":
                styleMap["bold"]=True
        if k == "color":
            styleMap["color"]=v
        if k == "text-decoration":
            if v in ['underline', 'line-through']:
                styleMap[v]=True
        if k == "font-style":
            if v=="italic":
                styleMap[v]=True
        if k == 'width':
            if v != "":
                styleMap['width']=v
        if k == "background-color" and v != False:
            styleMap['background-color']=v

    return styleMap

def applyFmt(tblStyle, trStyle,tdStyle, cell,ws):
    # resolve all the styles
    finalStyle=deepcopy(tblStyle)
    if finalStyle == None:
        finalStyle ={}
    for s in [trStyle,tdStyle]:
        if s==None:
            continue
        for k,v in s.iteritems():
            if v == False:
                continue
            finalStyle[k]=v
    font=Font()
    for k,v in finalStyle.iteritems():
        if k == "italic" and v!=False:
            font.i=True
        if k == "underline" and v!=False:
            font.u=Font.UNDERLINE_SINGLE
        if k == "line-through" and v!=False:
            font.strikethrough=True
        if k == "font_name" and v!=False:
            font.name=v
        if k=="bold" and v==True:
            font.bold=True
        if k=='width' and v != "" and v != False:
            c,r=coordinate_from_string(cell.coordinate)
            m=re.match("([\d\.]+)(\D+)",v)
            if m != None:
                w=m.group(1)
                units=m.group(2)
                if units == "in":
                    w=float(w)*12
            ws.column_dimensions[c].width=w
        if k == "color" and v != False:
            font.color = v[1:]
        if k == "background-color" and v != False:
            c=Color(v[1:])
            fill=PatternFill(patternType=fills.FILL_SOLID,fgColor=c)
            cell.fill = fill
            
    cell.font=font        

def trToExcel(ws,row,rowCount,tblStyle):
    trStyle = procStyle(row)
    offset=0
    for colCount,cell in zip(range(len(row.cells)),row.cells):
        rowspan,colspan, colMergeOffset=tdToExcel(ws,cell,colCount+offset,rowCount,trStyle,tblStyle)
        offset+=colMergeOffset

def tdToExcel(ws,wikicell,colCount, rowCount, trStyle, tblStyle): 
    colspan=0
    rowspan=0
    colMergeOffset=0
    tdStyle = procStyle(wikicell)
    cell=ws.cell(column=colCount+1,row=rowCount+1)
    while cell.coordinate in ws.merged_cells:
        colMergeOffset+=1
        cell=ws.cell(column=colCount+colMergeOffset+1,row=rowCount+1)

    cell.value=wikicell.text
    #replace <br/> with return characters
    cell.value=HTML_BR.sub("\n",cell.value)
    # if line contains multiple lines, then set the wrap style on
    # this means return character exists between texts
    if re.search(r".+\n.+", wikicell.text):
        cell.alignment = Alignment(wrapText=True)
        print "Wrap:",cell.coordinate

    if wikicell.attrs.has_key('colspan'):
        colspan+=int(wikicell.attrs['colspan']) -1
    if wikicell.attrs.has_key('rowspan'):
        rowspan=int(wikicell.attrs['rowspan']) -1
    if colspan != 0 or rowspan != 0:
        ws.merge_cells(start_row=rowCount+1,start_column=colCount+1,
                       end_row=rowCount+rowspan+1, end_column=colCount+colspan+1)
    
    applyFmt(tblStyle,trStyle,tdStyle,cell,ws)
    return rowspan, colspan, colMergeOffset
    


def wikiTblToExcel(wikiTblList):
    tblCount=0
    wb=Workbook()
    for tblInd,tbl in zip(range(len(wikiTblList)+1),wikiTblList):
        caption=tbl.caption
        if caption != None and caption != u"":
            shtName=captionToExcel(caption)
        else:
            shtName = "Sheet"+str(tblInd+1)

        ws=wb.create_sheet(index=tblInd+1, title=shtName)
        # set the new sheet as active
        wb.active=wb.get_index(ws)
        tblStyle=procStyle(tbl)
        for rowCount,row in zip(range(len(tbl.rows)),tbl.rows):
            trToExcel(ws,row,rowCount,tblStyle)
    return wb

class wikiToExcel():
    wb=None
    wikiContent=None
    def __init__(self,wikiContent=None,infile=None, ):
        """wikiContent is a string containing the wiki text
        or infile can be a filepath or a file-like object """

        if wikiContent != None:
            self.wikTblList= wikiTableParser(wikiContent)
            self.wikiContent=wikiContent
        elif infile != None:
            if isinstance(infile, file):
                self.wikiContent= infile.read()
            elif os.path.exists(infile):
                with open(infile,"r") as wikiFile:
                    self.wikiContent= wikiFile.read()
            else:
                raise ValueError("Input file or path is required: %s"%infile)
        else:
            raise ValueError("Insufficient arguments")
        self.wikiTblList= wikiTableParser(self.wikiContent)
        self.wb = wikiTblToExcel(self.wikiTblList)
    
    def getHTML(self):
        return self.wikiContent

    def getWorkBook(self):
        return self.wb
    
    def saveExcel(self,fileObj=None,fileName=None):
        """fileObj is a file like object
        fileName is path and name of file to write 
        one of the two should be present"""
        import sys
        if fileObj != None:
            try:
                self.wb.save(fileObj)
            except:
                print sys.exc_info()
                raise Exception("Could not save excel to: %s"%fileObj)
        elif fileName != None:
            with open(fileName,"w") as fileObj:
                try:
                    self.wb.save(fileObj)
                except:
                    print sys.exc_info()
                    raise Exception("Could not save excel to: %s"%fileName)
                
                
